from datetime import datetime
from hdbcli import dbapi
from openpyxl.styles import Border, Side, PatternFill, Alignment
from openpyxl import load_workbook
import pandas as pd

current_date = (datetime.now()).strftime("%d.%m.%Y_%H%M")


def get_data():
    # set the connection detalis
    connection = dbapi.connect(
        address="103.25.172.160",
        port=30015,
        database="RABYTE_PTE_LIVE",
        user="SYSTEM",
        password="Data1234",
    )
    print("🟢 Connection Done")
    query = f"""
    SELECT "LineNum" as "Item No",T1."U_UTL_BPREF" as "PO Number" , 'Invoice' as "Type" , T0."DocNum" as "Invoice Document No",
T0."CardCode" as "Customer Code",T0."CardName" as "Customer Name",T0."Address" as "Address 1",
 T0."Address2" as "Address 2", T3."StateB" as "State",  T3."CountyB" as "Country",
T3."ZipCodeB" as "Zip Code",T3."CityB" as "City",T1."LineNum" as "Lines Item No",T1."ItemCode" as "Item Code",
T1."Dscription" as "Item Name/Description",T1."unitMsr" as "UOM",
T1."Quantity" as "Qty",T1."Price" as "Sale Unit Price",T0."DocCur" as "Currency",'' as "Disty Book Cost",
'' as "Adjust Cost",T1."U_UTL_BPREF" as "Customer PO",
T4."U_RENESCMNO" as "SCM Code",T5."U_COO" as "COO-Company of Origin",T4."U_RENE_ABU" as "ABU/IIBU",
'' as "End Customer Code",
T0."ShipToCode" as "End Customer Name",
T0."Address2" as "End Address 2" ,

T3."StateS" as "End State",
T3."CountyS" as "End Country",  
T3."ZipCodeS" as "End Zip Code",
T3."CityS" as "End City",
'' as "SCM No of End Customer",
T1."U_SAMPLETYPE" as "Sample Type",
T1."U_FOC" as "FOC Sample",
'' as "S&D No",
'' as "S&D Qty",'' as "S&D Balance Qty",
'' as "S&D Price",
'' as "S&D Date From",
'' as "S&D Date To",
'' as "S&D Customer Code",
'' as "S&D Customer Name",
'' as "S&D Customer Name",
T4."U_OsramCustNo" as  "Osram End CustNo",
T0."U_Project" as "Project Name",
'' as "Project No",
T7."Name" "Make",
T6."Name" as "Make+Div"

FROM "RABYTE_PTE_LIVE".OINV T0

INNER JOIN "RABYTE_PTE_LIVE".INV1 T1 ON T1."DocEntry"=T0."DocEntry"
LEFT JOIN "RABYTE_PTE_LIVE".INV12 T3 ON T3."DocEntry"=T1."DocEntry"
LEFT JOIN "RABYTE_PTE_LIVE".OITM T2 ON T2."ItemCode"=T1."ItemCode"
LEFT JOIN "RABYTE_PTE_LIVE".OCRD T4 ON T4."CardCode"=T0."CardCode"
LEFT JOIN "RABYTE_PTE_LIVE".OITM T5 ON T5."ItemCode"=T1."ItemCode"
LEFT JOIN "RABYTE_PTE_LIVE"."@MAKE_FULL_2" T6 ON T6."Code"= T5."U_M_F_2"
LEFT JOIN "RABYTE_PTE_LIVE"."@MAKE_FULL_1" T7 ON T7."Code"= T5."U_MF_1"
WHERE T0."DocType"='I'

"""
    cursor2 = connection.cursor()
    cursor2.execute(query)
    data = cursor2.fetchall()
    column_names2 = [i[0] for i in cursor2.description]
    df2 = pd.DataFrame(data, columns=column_names2)

    # concatinating and saving in excel file
    df2.to_excel(
        f"files/output columns {current_date}.xlsx", index=False
    )
    wb = load_workbook(f"files/output columns {current_date}.xlsx")
    ws = wb.active
    border_style = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000"),
    )
    header_fill = PatternFill(
        start_color="1ec71e", end_color="1ec71e", fill_type="solid"
    )
    for cell in ws[1]:
        cell.fill = header_fill
        cell.border = border_style
        cell.alignment = Alignment(horizontal="left", vertical="bottom")
        column_letter = cell.column_letter
        ws.column_dimensions[column_letter].width = 15
    wb.save(f"files/output columns {current_date}.xlsx")
    print("🟡File Saved")

if __name__ == "__main__":
    print("✅✅Starting")
    get_data()
    print("🔴🔴Ending")
